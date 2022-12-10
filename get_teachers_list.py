from openpyxl import load_workbook


def get_names(path: str):
    wb = load_workbook(path)
    ws = wb.active
    teachers_list = set()

    for row in range(1, ws.max_row+1):
        for col in range(1, ws.max_column+1):
            cell = ws._get_cell(row, col)

            if cell.value != None:
                value = del_spaces(str(cell.value))
                value_ed = del_all_spaces(value)

                if 'п/г' in value:
                    sep = value.rfind('п/г')
                    text_1 = value[:sep-3]
                    text_2 = value[sep-2:]

                    if text_1[-3] == '.' and (text_1[-1] == '.' or text_1[-1] == ','):
                        text = rename(text_1).split()
                        teachers_list.add(f'{text[-2]} {text[-1]}')

                    if text_2[-3] == '.' and (text_2[-1] == '.' or text_2[-1] == ','):
                        text = rename(text_2).split()
                        teachers_list.add(f'{text[-2]} {text[-1]}')
                
                elif '1 н' in value_ed or '2 н' in value_ed or '1н' in value_ed or '2н' in value_ed:
                    values = del_spaces(value).split('2 н')
                    if len(values) == 1:
                        values = del_spaces(value).split('2н')
                    
                    text_1 = del_spaces(values[0])
                    text_2 = del_spaces(values[-1])

                    if text_1[-3] == '.' and (text_1[-1] == '.' or text_1[-1] == ',') and text_1[-2] != ' ':
                        text = rename(text_1).split()
                        teachers_list.add(f'{text[-2]} {text[-1]}')
                
                    if text_2[-3] == '.' and (text_2[-1] == '.' or text_2[-1] == ',') and text_2[-2] != ' ':
                        text = rename(text_2).split()
                        teachers_list.add(f'{text[-2]} {text[-1]}')
                
                elif len(value) >= 3:
                    if value[-3] == '.' and (value[-1] == '.' or value[-1] == ','):
                        text = rename(value).split()
                        teachers_list.add(f'{text[-2]} {text[-1]}')
                     
    return sorted(list(teachers_list))


def del_spaces(string: str) -> str:
    if string != None:
        return ' '.join([str(i) for i in string.split()])

def del_all_spaces(string: str) -> str:
    if string != None:
        return string.replace(' ', '')

def rename(string: str):
    if string[-1] == '.' and string[-3] == '.':
        return string
    elif string[-1] == ',' and string[-3] == '.':
        string = list(string)
        string[-1] = '.'
        return ''.join([s for s in string]) 


print(get_names(path=r"resources\tables\RASPISANIE.xlsx"))