from openpyxl import load_workbook
from calendar import monthrange
from teachers_list import teachers_list
from glob import glob

from data import days_strs, offset
from config import config
from Date import Date
from ExcelFile import ExcelFile


import time
import shutil

class Table():
    def __init__(self, config: dict) -> None:
        self._table_path = config['table_path']
        self._timetable_path = config['timetable_path']
        self._changes_folder_path = config['changes_folder_path']
        self._folder_to_save = config['folder_to_save']
        self._templates_folder = config['templates_folder']

        self._create_table()

    def fill_table(self) -> None:
        wb = load_workbook(self.table)
        ws = wb.active

        wb_tt = load_workbook(self._timetable_path)

        last_full_column = current_str = 2

        for teacher in teachers_list:
            ws._get_cell(row=1, column=last_full_column+1).value = teacher
            last_full_column += 1

        for date in self._get_days_dict():
            ws[f'A{current_str}'].value = f'{date}\n{Date(date).get_day_name()}'
            
            if 'суббота' in str(ws[f'A{current_str}'].value) or 'воскресенье' in str(ws[f'A{current_str}'].value):
                wb.save(self.table)
                current_str += 6
                continue

            for teacher_num in range(len(teachers_list)):
                teacher = teachers_list[teacher_num].split()[0]

                for ws_tt in wb_tt:
                    for row in range(days_strs[Date(date).get_day_name()], days_strs[Date(Date(date).increase_day(1)).get_day_name()]):
                        for col in range(1, ws_tt.max_column+1):
                            cell = ws_tt._get_cell(row=row, column=col)
                            value = str(cell.value)

                            if teacher in value.split():
                                time = self._del_all_spaces(ws_tt._get_cell(row=row, column=2).value)
                                group = self._del_ext_spaces(ws_tt._get_cell(row=6, column=cell.column).value)
                                current_cell = ws._get_cell(row=current_str+offset[time], column=teacher_num+3)
                            
                                if current_cell.value == None:
                                    current_cell.value = f'{group}\n'
                                else:
                                    current_cell.value = current_cell.value+f'{group}\n'
                else:
                    wb.save(self.table)

                if self._get_days_dict()[date] != None:
                    wb_ch = load_workbook(f'{self._changes_folder_path}\{date}.xlsx')
                    ws_ch = wb_ch.active
                    for row in ws_ch.rows:
                        if teacher in str(row[2].value).replace('\n', ' ').split() and str(row[2].value) != str(row[3].value) and row[0].value != None:
                            group = str(row[1].value)
                            lesson_num = int(row[0].value)
                            group_list = ws._get_cell(row=current_str+lesson_num-1, column=teacher_num+3).value

                            if group_list != None:
                                group_list = group_list.split('\n')
                            else:
                                group_list = ''
                            
                            if group in group_list:
                                text = '\n'.join(gr for gr in group_list if gr != group)
                                ws._get_cell(row=current_str+lesson_num-1, column=teacher_num+3).value = f'{text}\n'


                        elif teacher in str(row[3].value).replace('\n', ' ').split() and str(row[2].value) != str(row[3].value) and row[0].value != None:
                            group = str(row[1].value)
                            lesson_num = int(row[0].value)
                            group_list = ws._get_cell(row=current_str+lesson_num-1, column=teacher_num+3).value

                            if group_list != None:
                                ws._get_cell(row=current_str+lesson_num-1, column=teacher_num+3).value = f'{group_list}{group}\n'
                            else:
                                ws._get_cell(row=current_str+lesson_num-1, column=teacher_num+3).value = f'{group}\n'
                    wb_ch.close()
                    wb.save(self.table)

            current_str += 6
        
        wb.save(self.table)



    def _create_table(self) -> None:
        shutil.copy2(src=self._select_tamplate(), dst=f'{self._folder_to_save}/Табель.xlsx')
        self.table = f'{self._folder_to_save}/Табель.xlsx'

    def _select_tamplate(self) -> str: #template path
        return f'{self._templates_folder}/base{self._get_days_amount()}.xlsx'

    def _get_days_dict(self) -> dict[str, str | None]:
        self._convert_xls_to_xlsx()

        days_dict = {}
        files_list = glob(f'{self._changes_folder_path}/*.xlsx')
        
        date = '.'.join(self._get_date_from_file().split('.')[1:])

        for day in range(1, self._get_days_amount()+1):
            if day < 10:
                current_date = f'0{day}.{date}'
            else:
                current_date = f'{day}.{date}'
            
            for file in files_list:
                if current_date in  file:
                    days_dict[current_date] = file
                    break
                else:
                    days_dict[current_date] = None

        return days_dict
    
    def _convert_xls_to_xlsx(self) -> None:
        files_list = glob(f'{self._changes_folder_path}/*.xls')
        for file in files_list:
            ExcelFile(file_path=file).convert_xls_to_xlsx()
    
    def _get_days_amount(self) -> int:
        date = self._get_date_from_file().split('.')
        days_amount = monthrange(year=int(date[-1]), month=int(date[1]))[-1]
        return days_amount

    def _get_date_from_file(self) -> str:
        return ExcelFile(glob(f'{self._changes_folder_path}/*.xlsx')[0]).get_name(extension=False)
    
    def _del_all_spaces(self, text: str) -> str:
        return text.replace(' ', '')
    
    def _del_ext_spaces(self, text: str) -> str:
        return ' '.join(text.split())

st = time.time()
table = Table(config=config)
#print(table._get_days_dict())
print(table.fill_table())
print(time.time()-st)
#ex = ExcelFile(r'resources\changes\замены\ЯНВАРЬ\31.01.2023.xls')
#print(ex.convert_xls_to_xlsx())